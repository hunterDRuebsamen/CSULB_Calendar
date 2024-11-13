from fastapi import FastAPI
from pydantic import BaseModel
import uvicorn
from .classes.month import CalMonth
from .classes.year import CalYear, Calendar_Input
from fastapi.responses import StreamingResponse
import base64
import io
from io import BytesIO
import time
import itertools
import logging
import openpyxl
from openpyxl.styles import PatternFill
from app.classes.year import DayType

app = FastAPI()

class Month(BaseModel):
    month: int = 1
    day_colors: list | None = ['red', 'blue']

class Base_Input(BaseModel):
    month: int
    day: int 
    year: int
    
@app.post("/calendar/test")
async def testThree(req: Month):
    return {
                "message": "hello world",
                "day_colors": req.day_colors
            }

@app.post("/calendar/table")
async def testTwo(req: Month):
    year = CalYear()
    awd_fall = [12, 21, 22, 19, 19]
    id_fall = [6, 21, 22, 16, 9]
    months = ['Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    img = year.create_months_table(awd=awd_fall, id=id_fall, months=months)
    
    img_bytes_io = io.BytesIO()
    img.save(img_bytes_io, format='PNG')
    
    img_bytes_io.seek(0)

    return StreamingResponse(img_bytes_io, media_type="image/png")

@app.post("/calendar/month")
async def testOne(req: Calendar_Input):
    cmonth = CalMonth(req.year, req.month)
    img = cmonth.draw(req.width)
    img_bytes_io = io.BytesIO()
    img.save(img_bytes_io, format='PNG')
    img_bytes_io.seek(0)
    
    return StreamingResponse(img_bytes_io, media_type="image/png")

@app.post("/calendar/build_year")
async def build_year(req: Calendar_Input):    
    start = time.perf_counter()
    calyear = CalYear(req)
    if calyear.valid:
        result_image = calyear.gen_schedule()
    
        img_bytes_io = io.BytesIO()
        result_image.save(img_bytes_io, format='PNG')
        
        img_bytes_io.seek(0)
        
        end = time.perf_counter()
        print("Elapsed (time) = {}s".format((end - start)))
        return StreamingResponse(img_bytes_io, media_type="image/png")
    return None

@app.post("/calendar/build_years_test")
def build_years_request2(req: Calendar_Input):    
    result = []
    calyear = CalYear(req)
    cnt = 0

    awd = 172
    id = 148
    convo_day = 3
    winter_sess = 13

    print(f"\nGenerating calendar {cnt}")
    start = time.perf_counter()
    result_image, a = calyear.gen_schedule(awd, id, convo_day, winter_sess)
    cnt += 1
    if result_image == None:
        pass
    else:
        img_bytes_io = io.BytesIO()
        result_image.save(img_bytes_io, format='PNG')
        img_str = base64.b64encode(img_bytes_io.getvalue()).decode("utf-8")
        returndict = {"image": img_str}
        result.append(returndict)
    end = time.perf_counter()
    print("Elapsed (time) = {}s".format((end - start)))
        
    return result

@app.post("/calendar/build_years_sequential")
def build_years_request2(req: Calendar_Input):    
    result = []
    calyear = CalYear(req)
    result_dict = {}
    cnt = 0
    for awd in range(170, 181):
        for id in range(145, 150):
            for convo_day in range(2, 6):
                for winter_sess in range(12, 16):
                    print(f"\nGenerating calendar {cnt}")
                    start = time.perf_counter()
                    result_image, md5_hash = calyear.gen_schedule(awd, id, convo_day, winter_sess)
                    if md5_hash in result_dict:
                        result_image = None
                        print("Duplicate calendar, discarded")
                    else:
                        result_dict[md5_hash] = True
                    cnt += 1
                    if result_image == None:
                        pass
                    else:
                        img_bytes_io = io.BytesIO()
                        result_image.save(img_bytes_io, format='PNG')
                        img_str = base64.b64encode(img_bytes_io.getvalue()).decode("utf-8")
                        returndict = {"image": img_str}
                        result.append(returndict)
                    end = time.perf_counter()
                    print("Elapsed (time) = {}s".format((end - start)))
        
    return result

@app.post("/calendar/build_years")
def build_years_request(req: Calendar_Input):
    calyear = CalYear(req)
    result_dict = {}

    args_list = [
        (i, calyear, awd, id, convo_day, winter_sess)
        for i, (awd, id, convo_day, winter_sess) in enumerate(
            itertools.product(range(170, 181), range(145, 150), range(2, 6), range(12, 16))
        )
    ]

    results = []
    for args in args_list:
        idx, calyear, awd, id, convo_day, winter_sess = args
        print(f"Processing calendar {idx}")
        result_image, md5_hash = calyear.gen_schedule(awd, id, convo_day, winter_sess)
        if md5_hash in result_dict:
            result_image = None
            print("Duplicate calendar, discarded")
        else:
            result_dict[md5_hash] = True

        if result_image is None:
            results.append({"image": None})
        else:
            img_bytes_io = io.BytesIO()
            result_image.save(img_bytes_io, format='PNG')
            img_str = base64.b64encode(img_bytes_io.getvalue()).decode("utf-8")
            results.append({"image": img_str})

        if len(results) % 10 == 0:
            logging.debug(f"Processed {len(results)} calendars")

    return results

def generate_colored_excel_calendar(calyear):
    wb = openpyxl.Workbook()
    
    day_type_colors = {
        DayType.AWD: "CCEFFF",  # Light Blue
        DayType.ID: "EFEF95",  # Yellow
        DayType.CONVOCATION: "50E3C2",  # Greenish
        DayType.FINALS: "F8CBAD",  # Peach
        DayType.NO_CLASS_CAMPUS_OPEN: "AE76C7",  # Purple
        DayType.COMMENCEMENT: "BD10E0",  # Magenta
        DayType.SUMMER_SESSION: "C5E0B4",  # Light Green
        DayType.WINTER_SESSION: "CFCFCF",  # Gray
        DayType.NONE: "FFFFFF",  # White
        DayType.HOLIDAY: "F8CBAD",  # Add a color for holiday
        DayType.VOID: "D9D9D9",  # Light Gray
    }

    print("Day Type to Color Mapping in generate_colored_excel_calendar:")
    for day_type, color in day_type_colors.items():
        print(f"{day_type}: {color}")

    for i, calmonth in enumerate(calyear.months):  # Assuming calyear.months holds months
        ws = wb.create_sheet(title=f"{calmonth.get_title()}")

        day_names = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
        for col, day_name in enumerate(day_names, start=1):
            ws.cell(row=1, column=col, value=day_name)

        for row, week in enumerate(calmonth.cal, start=2):  
            for col, day in enumerate(week, start=1):
                if day == 0:
                    continue

                cell = ws.cell(row=row, column=col, value=day if day != 0 else "")
                
                day_type = calyear.get_day_type(calmonth.year, calmonth.month, day)  # Get day type
                if day_type and day_type != "None":
                    print(f"Day: {day}, Type: {day_type}, Color: {day_type_colors[day_type]}")
                    fill_color = PatternFill(start_color=day_type_colors[day_type], end_color=day_type_colors[day_type], fill_type="solid")
                    cell.fill = fill_color

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    return excel_stream

# FastAPI route for downloading the Excel file
@app.post("/calendar/download_excel_colored")
async def download_calendar_excel_colored(req: Calendar_Input):
    logging.debug("Received request for download_excel_colored with data:")
    logging.debug(req.dict())
    calyear = CalYear(req)
    if calyear.valid:
        excel_file = generate_colored_excel_calendar(calyear)
        headers = {
            "Content-Disposition": "attachment; filename=calendar_colored.xlsx"
        }
        return StreamingResponse(
            excel_file,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
    else:
        logging.error("Invalid Calendar Input")
        return {"message": "Invalid Calendar"}

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)

